"""Excel export of a :class:`PortfolioReport`, using openpyxl.

Monetary and quantity values are converted to ``float`` before writing:
the XLSX file format itself has no ``Decimal`` type — every numeric cell
is stored as a double internally — so converting here is honest about
what the file format can actually represent, rather than pretending
openpyxl preserves ``Decimal`` precision end to end.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from finapp.application.dto import PortfolioDividendIncome, PortfolioReport
from finapp.application.ports import PortfolioReportExporter

_HEADER_FONT = Font(bold=True)
_POSITION_HEADERS = (
    "Symbol",
    "Quantity",
    "Avg Cost",
    "Price",
    "Book Cost",
    "Market Value",
    "Unrealized P&L",
)
_DIVIDEND_HEADERS = ("Symbol", "Quantity Held", "Amount/Share", "Pay Date", "Total Income")


class ExcelPortfolioReportExporter(PortfolioReportExporter):
    """Renders a :class:`PortfolioReport` as an .xlsx workbook: a Summary
    sheet, a Positions sheet, and (if dividend income was included in the
    report) a Dividends sheet.
    """

    def export(self, report: PortfolioReport, output_path: Path) -> Path:
        workbook = Workbook()
        summary_sheet = workbook.active
        assert summary_sheet is not None  # a freshly-created Workbook always has one
        self._write_summary_sheet(summary_sheet, report)
        self._write_positions_sheet(workbook.create_sheet("Positions"), report)
        if report.dividend_income is not None:
            self._write_dividends_sheet(workbook.create_sheet("Dividends"), report.dividend_income)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path

    @staticmethod
    def _write_summary_sheet(sheet: Worksheet, report: PortfolioReport) -> None:
        sheet.title = "Summary"
        valuation = report.valuation
        rows = (
            ("Portfolio", report.portfolio_name),
            ("Generated at", report.generated_at.isoformat()),
            ("Total book cost", str(valuation.base_currency_total_book_cost)),
            ("Total market value", str(valuation.base_currency_total_market_value)),
            ("Total unrealized P&L", str(valuation.base_currency_total_unrealized_pnl)),
        )
        for row_index, (label, value) in enumerate(rows, start=1):
            sheet.cell(row=row_index, column=1, value=label).font = _HEADER_FONT
            sheet.cell(row=row_index, column=2, value=value)
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 32

    @staticmethod
    def _write_positions_sheet(sheet: Worksheet, report: PortfolioReport) -> None:
        for column_index, header in enumerate(_POSITION_HEADERS, start=1):
            sheet.cell(row=1, column=column_index, value=header).font = _HEADER_FONT

        for row_index, position in enumerate(report.valuation.positions, start=2):
            sheet.cell(row=row_index, column=1, value=position.symbol)
            sheet.cell(row=row_index, column=2, value=float(position.quantity))
            sheet.cell(row=row_index, column=3, value=float(position.average_cost.amount))
            sheet.cell(row=row_index, column=4, value=float(position.market_price.amount))
            sheet.cell(row=row_index, column=5, value=float(position.book_cost.amount))
            sheet.cell(row=row_index, column=6, value=float(position.market_value.amount))
            sheet.cell(row=row_index, column=7, value=float(position.unrealized_pnl.amount))

        for column_letter in ("A", "B", "C", "D", "E", "F", "G"):
            sheet.column_dimensions[column_letter].width = 16

    @staticmethod
    def _write_dividends_sheet(sheet: Worksheet, dividend_income: PortfolioDividendIncome) -> None:
        for column_index, header in enumerate(_DIVIDEND_HEADERS, start=1):
            sheet.cell(row=1, column=column_index, value=header).font = _HEADER_FONT

        row_index = 2
        for income in dividend_income.incomes:
            sheet.cell(row=row_index, column=1, value=income.instrument.symbol)
            sheet.cell(row=row_index, column=2, value=float(income.quantity_held))
            sheet.cell(row=row_index, column=3, value=float(income.dividend.amount_per_share.amount))
            sheet.cell(row=row_index, column=4, value=income.dividend.pay_date.isoformat())
            sheet.cell(row=row_index, column=5, value=float(income.total_income.amount))
            row_index += 1

        sheet.cell(row=row_index, column=1, value="Total").font = _HEADER_FONT
        sheet.cell(
            row=row_index, column=5, value=float(dividend_income.base_currency_total_income.amount)
        ).font = _HEADER_FONT

        for column_letter in ("A", "B", "C", "D", "E"):
            sheet.column_dimensions[column_letter].width = 16
