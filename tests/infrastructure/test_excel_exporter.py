from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from finapp.application.dto import (
    DividendIncome,
    PortfolioDividendIncome,
    PortfolioReport,
    PortfolioValuation,
    PositionValuation,
)
from finapp.domain.entities.instrument import Instrument
from finapp.domain.value_objects.dividend import Dividend
from finapp.domain.value_objects.enums import AssetType, Currency
from finapp.domain.value_objects.money import Money
from finapp.infrastructure.reporting.excel_exporter import ExcelPortfolioReportExporter


def _tlv() -> Instrument:
    return Instrument(
        symbol="TLV", name="Banca Transilvania", currency=Currency.RON, asset_type=AssetType.EQUITY
    )


def _valuation() -> PortfolioValuation:
    position = PositionValuation(
        symbol="TLV",
        quantity=Decimal("100"),
        average_cost=Money(amount=Decimal("4.00"), currency=Currency.RON),
        market_price=Money(amount=Decimal("4.50"), currency=Currency.RON),
        book_cost=Money(amount=Decimal("400.00"), currency=Currency.RON),
        market_value=Money(amount=Decimal("450.00"), currency=Currency.RON),
        unrealized_pnl=Money(amount=Decimal("50.00"), currency=Currency.RON),
    )
    return PortfolioValuation(
        portfolio_name="Retirement",
        base_currency_total_book_cost=Money(amount=Decimal("400.00"), currency=Currency.RON),
        base_currency_total_market_value=Money(amount=Decimal("450.00"), currency=Currency.RON),
        base_currency_total_unrealized_pnl=Money(amount=Decimal("50.00"), currency=Currency.RON),
        positions=(position,),
    )


def _dividend_income() -> PortfolioDividendIncome:
    income = DividendIncome(
        instrument=_tlv(),
        quantity_held=Decimal("100"),
        dividend=Dividend(
            symbol="TLV",
            amount_per_share=Money(amount=Decimal("0.25"), currency=Currency.RON),
            pay_date=date(2026, 6, 14),
        ),
        total_income=Money(amount=Decimal("25.00"), currency=Currency.RON),
    )
    return PortfolioDividendIncome(
        portfolio_name="Retirement",
        base_currency_total_income=Money(amount=Decimal("25.00"), currency=Currency.RON),
        incomes=(income,),
    )


def test_exports_summary_and_positions_sheets(tmp_path: Path) -> None:
    report = PortfolioReport(
        portfolio_name="Retirement",
        generated_at=datetime(2026, 8, 27, tzinfo=UTC),
        valuation=_valuation(),
    )
    output_path = tmp_path / "report.xlsx"

    result_path = ExcelPortfolioReportExporter().export(report, output_path)

    assert result_path == output_path
    assert output_path.exists()

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Summary", "Positions"]

    summary = workbook["Summary"]
    assert summary.cell(row=1, column=1).value == "Portfolio"
    assert summary.cell(row=1, column=2).value == "Retirement"

    positions = workbook["Positions"]
    assert positions.cell(row=1, column=1).value == "Symbol"
    assert positions.cell(row=2, column=1).value == "TLV"
    assert positions.cell(row=2, column=2).value == 100.0
    assert positions.cell(row=2, column=6).value == 450.0


def test_includes_dividends_sheet_when_present(tmp_path: Path) -> None:
    report = PortfolioReport(
        portfolio_name="Retirement",
        generated_at=datetime(2026, 8, 27, tzinfo=UTC),
        valuation=_valuation(),
        dividend_income=_dividend_income(),
    )
    output_path = tmp_path / "report.xlsx"

    ExcelPortfolioReportExporter().export(report, output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Summary", "Positions", "Dividends"]

    dividends = workbook["Dividends"]
    assert dividends.cell(row=1, column=1).value == "Symbol"
    assert dividends.cell(row=2, column=1).value == "TLV"
    assert dividends.cell(row=2, column=5).value == 25.0
    assert dividends.cell(row=3, column=1).value == "Total"
    assert dividends.cell(row=3, column=5).value == 25.0


def test_creates_parent_directories(tmp_path: Path) -> None:
    report = PortfolioReport(
        portfolio_name="Retirement",
        generated_at=datetime(2026, 8, 27, tzinfo=UTC),
        valuation=_valuation(),
    )
    output_path = tmp_path / "nested" / "dir" / "report.xlsx"

    ExcelPortfolioReportExporter().export(report, output_path)

    assert output_path.exists()
